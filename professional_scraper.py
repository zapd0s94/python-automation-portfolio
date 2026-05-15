import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import logging
from datetime import datetime
from pathlib import Path
import random

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("scraper_log.txt", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

CONFIG = {
    "base_url": "https://books.toscrape.com/catalogue/page-{}.html",
    "max_pages": 50,
    "delay_min": 1.0,
    "delay_max": 2.5,
    "max_retries": 3,
    "output_dir": "outputs",
    "user_agents": [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/605.1.15",
    ]
}

def get_headers():
    return {"User-Agent": random.choice(CONFIG["user_agents"])}

def fetch_page(url, retry=0):
    try:
        r = requests.get(url, headers=get_headers(), timeout=15)
        r.raise_for_status()
        log.info(f"OK [{r.status_code}] → {url}")
        return r
    except Exception as e:
        log.warning(f"Error: {e}")
        if retry < CONFIG["max_retries"]:
            time.sleep(5 * (retry + 1))
            return fetch_page(url, retry + 1)
        return None

def parse_products(soup):
    products = []
    items = soup.select("article.product_pod")
    log.info(f"Items encontrados en página: {len(items)}")

    for item in items:
        try:
            name = item.select_one("h3 a")["title"]
            price = float(
                item.select_one("p.price_color")
                .text.strip()
                .encode("ascii", "ignore")
                .decode()
                .replace("£", "")
                .replace("Â", "")
                .strip()
            )
            rating_map = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}
            rating_num = rating_map.get(item.select_one("p.star-rating")["class"][1], 0)
            rating_stars = "★" * rating_num + "☆" * (5 - rating_num)
            availability = item.select_one("p.availability").text.strip()
            href = item.select_one("h3 a")["href"].replace("../", "")
            if not href.startswith("http"):
                href = "https://books.toscrape.com/catalogue/" + href

            products.append({
                "nombre": name,
                "precio_GBP": price,
                "precio_USD": round(price * 1.27, 2),
                "rating": rating_stars,
                "rating_num": rating_num,
                "disponibilidad": availability,
                "url": href,
                "fecha_scraping": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
        except Exception as e:
            log.warning(f"Error parseando item: {e}")
    return products

def save_results(data):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    Path(CONFIG["output_dir"]).mkdir(exist_ok=True)
    df = pd.DataFrame(data)
    df = df.drop_duplicates(subset=["nombre"])
    df = df.sort_values("precio_GBP").reset_index(drop=True)

    export_cols = ["nombre", "precio_GBP", "precio_USD", "rating", "disponibilidad", "url", "fecha_scraping"]
    df_export = df[export_cols].copy()

    filename = f"outputs/scraping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    DARK_BLUE   = "0D2137"
    MID_BLUE    = "1565C0"
    ALT_ROW     = "F0F7FF"
    WHITE       = "FFFFFF"
    GOLD        = "F9A825"
    GRAY_BORDER = "CFD8DC"
    SUMMARY_BG  = "E8EAF6"
    SUMMARY_HDR = "3949AB"

    def thin_border():
        s = Side(style="thin", color=GRAY_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Productos", index=False, startrow=3)

        # ── HOJA PRODUCTOS ──────────────────────────────────────────
        ws = writer.sheets["Productos"]
        ws.sheet_view.showGridLines = False

        # Fila 1 — Banner título
        ws.merge_cells("A1:G1")
        banner = ws["A1"]
        banner.value = f"🕷️  PROFESSIONAL WEB SCRAPER  ·  {len(df_export)} productos  ·  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        banner.fill = PatternFill("solid", fgColor=DARK_BLUE)
        banner.font = Font(bold=True, color=GOLD, size=13, name="Arial")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Fila 2 — Sub-banner métricas
        ws.merge_cells("A2:G2")
        min_p = df["precio_GBP"].min()
        max_p = df["precio_GBP"].max()
        avg_p = df["precio_GBP"].mean()
        avg_r = df["rating_num"].mean()
        sub = ws["A2"]
        sub.value = (f"💰 Precio mín: £{min_p:.2f}  |  Precio máx: £{max_p:.2f}  |  "
                     f"Promedio: £{avg_p:.2f}  |  ⭐ Rating promedio: {avg_r:.1f}/5  |  "
                     f"✅ In Stock: {(df['disponibilidad']=='In stock').sum()}")
        sub.fill = PatternFill("solid", fgColor=MID_BLUE)
        sub.font = Font(bold=False, color=WHITE, size=10, name="Arial")
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Fila 3 — separadora
        ws.merge_cells("A3:G3")
        ws["A3"].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws.row_dimensions[3].height = 6

        # Fila 4 — Headers
        headers_display = ["📖 Nombre", "💷 Precio (GBP)", "💵 Precio (USD)", "⭐ Rating", "📦 Disponibilidad", "🔗 URL", "🕐 Fecha Scraping"]
        for i, h in enumerate(headers_display, 1):
            cell = ws.cell(row=4, column=i)
            cell.value = h
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
            cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[4].height = 28

        # Filas de datos
        for row_idx in range(5, len(df_export) + 5):
            fill_color = ALT_ROW if row_idx % 2 == 0 else WHITE
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = thin_border()
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col_idx in (2, 3, 4, 5, 7) else "left"
                )

            # Precio USD verde si bajo promedio
            precio_cell = ws.cell(row=row_idx, column=3)
            try:
                val = float(precio_cell.value) if precio_cell.value else 0
                if val < avg_p * 1.27:
                    precio_cell.font = Font(name="Arial", size=10, color="1B5E20", bold=True)
            except:
                pass

            # Disponibilidad coloreada
            disp_cell = ws.cell(row=row_idx, column=5)
            if disp_cell.value == "In stock":
                disp_cell.font = Font(name="Arial", size=10, color="1B5E20", bold=True)
            else:
                disp_cell.font = Font(name="Arial", size=10, color="B71C1C", bold=True)

        # Anchos de columna
        col_widths = [48, 16, 16, 14, 16, 58, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A5"

        # Fila totales
        total_row = len(df_export) + 5
        ws.merge_cells(f"A{total_row}:G{total_row}")
        tc = ws[f"A{total_row}"]
        tc.value = f"TOTAL: {len(df_export)} productos únicos extraídos  ·  github.com/zapd0s94/python-automation-portfolio"
        tc.fill = PatternFill("solid", fgColor=DARK_BLUE)
        tc.font = Font(bold=True, color=GOLD, size=10, name="Arial")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[total_row].height = 22

        # ── HOJA ESTADÍSTICAS ────────────────────────────────────────
        wb = writer.book
        ws2 = wb.create_sheet("Estadisticas")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:C1")
        b2 = ws2["A1"]
        b2.value = "📊  REPORTE DE ESTADÍSTICAS — WEB SCRAPER"
        b2.fill = PatternFill("solid", fgColor=DARK_BLUE)
        b2.font = Font(bold=True, color=GOLD, size=13, name="Arial")
        b2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 32

        by_rating = df.groupby("rating_num").size()
        stats_rows = [
            ("", "", ""),
            ("📋  RESUMEN GENERAL", "", ""),
            ("Métrica", "Valor", "Detalle"),
            ("Total de productos", len(df), "únicos, sin duplicados"),
            ("Páginas scrapeadas", CONFIG["max_pages"], "páginas procesadas"),
            ("Precio mínimo (GBP)", f"£{df['precio_GBP'].min():.2f}", df[df['precio_GBP'] == df['precio_GBP'].min()]['nombre'].iloc[0][:40]),
            ("Precio máximo (GBP)", f"£{df['precio_GBP'].max():.2f}", df[df['precio_GBP'] == df['precio_GBP'].max()]['nombre'].iloc[0][:40]),
            ("Precio promedio (GBP)", f"£{df['precio_GBP'].mean():.2f}", ""),
            ("Precio mediana (GBP)", f"£{df['precio_GBP'].median():.2f}", ""),
            ("Productos en stock", int((df['disponibilidad'] == 'In stock').sum()), f"{(df['disponibilidad']=='In stock').mean()*100:.1f}% del total"),
            ("Rating promedio", f"{df['rating_num'].mean():.2f} / 5", "★" * round(df['rating_num'].mean())),
            ("", "", ""),
            ("⭐  DISTRIBUCIÓN POR RATING", "", ""),
            ("Rating", "Cantidad", "Porcentaje"),
        ]
        for r in range(1, 6):
            count = int(by_rating.get(r, 0))
            pct = count / len(df) * 100
            stats_rows.append((f"{'★'*r}{'☆'*(5-r)}  ({r}/5)", count, f"{pct:.1f}%"))

        for i, row_data in enumerate(stats_rows, start=2):
            for j, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=i, column=j, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if j > 1 else "left"
                )
                if row_data[0] not in ("", "📋  RESUMEN GENERAL", "⭐  DISTRIBUCIÓN POR RATING"):
                    cell.border = thin_border()

        for i, row_data in enumerate(stats_rows, start=2):
            if row_data[0] in ("📋  RESUMEN GENERAL", "⭐  DISTRIBUCIÓN POR RATING"):
                ws2.merge_cells(f"A{i}:C{i}")
                cell = ws2.cell(row=i, column=1)
                cell.fill = PatternFill("solid", fgColor=MID_BLUE)
                cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws2.row_dimensions[i].height = 24
            elif row_data[0] in ("Métrica", "Rating"):
                for j in range(1, 4):
                    c = ws2.cell(row=i, column=j)
                    c.fill = PatternFill("solid", fgColor=SUMMARY_HDR)
                    c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                ws2.row_dimensions[i].height = 22
            elif i % 2 == 0 and row_data[0] != "":
                for j in range(1, 4):
                    c = ws2.cell(row=i, column=j)
                    if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                        c.fill = PatternFill("solid", fgColor=SUMMARY_BG)

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 45

    log.info(f"✓ Guardado: {filename} — {len(df)} productos únicos")
    return filename, df

def run_scraper():
    log.info("=" * 50)
    log.info("SCRAPER PROFESIONAL — INICIANDO")
    log.info("=" * 50)

    all_products = []

    for page in range(1, CONFIG["max_pages"] + 1):
        url = CONFIG["base_url"].format(page)
        log.info(f"Página {page}: {url}")

        response = fetch_page(url)
        if response is None:
            log.error(f"No se pudo acceder a página {page} — deteniendo")
            break

        soup = BeautifulSoup(response.text, "html.parser")
        products = parse_products(soup)
        all_products.extend(products)
        log.info(f"Total acumulado: {len(all_products)}")

        if not soup.select_one("li.next"):
            log.info("Última página alcanzada")
            break

        time.sleep(random.uniform(CONFIG["delay_min"], CONFIG["delay_max"]))

    if all_products:
        filename, df = save_results(all_products)
        log.info("=" * 50)
        log.info(f"COMPLETADO: {len(df)} productos")
        log.info(f"Precio mín: £{df['precio_GBP'].min()} | máx: £{df['precio_GBP'].max()}")
        log.info("=" * 50)
    else:
        log.error("Sin datos")

if __name__ == "__main__":
    run_scraper()